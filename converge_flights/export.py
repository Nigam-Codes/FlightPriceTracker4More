"""Excel (.xlsx) export with openpyxl, including a charts dashboard.

Three tabs:
    * **Fares** — one row per traveler × window best offer.
    * **Comparison** — windows ranked by group total, cheapest origin/provider
      per traveler, and savings versus the most expensive window.
    * **Dashboard** — a grouped bar chart of group cost by window and a line
      chart of each traveler's cost trend across windows.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from converge_flights.models import ComparisonReport


def _autosize(ws: Worksheet) -> None:
    """Roughly fit column widths to their contents."""
    for column_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=8,
        )
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)


def _fares_tab(wb: Workbook, report: ComparisonReport) -> None:
    ws = wb.active
    ws.title = "Fares"
    ws.append(
        [
            "Traveler",
            "Depart",
            "Return",
            "Provider",
            "Origin",
            "Destination",
            "Price",
            "Currency",
            "Max stops",
            "Outbound h",
            "Inbound h",
            "Carrier",
            "Qualifying offers",
        ]
    )
    for window in report.ranked():
        for name, result in window.per_traveler.items():
            offer = result.best_offer
            if offer is None:
                ws.append(
                    [
                        name,
                        window.window.depart.isoformat(),
                        window.window.return_.isoformat(),
                        "—",
                        "—",
                        report.currency,
                        None,
                        report.currency,
                        None,
                        None,
                        None,
                        "NO QUALIFYING OPTION",
                        0,
                    ]
                )
                continue
            ws.append(
                [
                    name,
                    window.window.depart.isoformat(),
                    window.window.return_.isoformat(),
                    offer.provider,
                    offer.origin,
                    offer.destination,
                    float(offer.price),
                    offer.currency,
                    offer.max_stops,
                    round(offer.outbound.duration_hours, 2),
                    round(offer.inbound.duration_hours, 2),
                    offer.carrier,
                    result.considered,
                ]
            )
    _autosize(ws)


def _comparison_tab(
    wb: Workbook, report: ComparisonReport, traveler_names: list[str]
) -> Worksheet:
    ws = wb.create_sheet("Comparison")
    baseline = report.most_expensive_total()
    header = ["Rank", "Depart", "Return", "Group total"]
    for name in traveler_names:
        header.append(f"{name} price")
        header.append(f"{name} won (provider/origin)")
    header += ["Savings vs. worst", "Complete", "Missing"]
    ws.append(header)

    for rank, window in enumerate(report.ranked(), start=1):
        total = window.group_total
        row: list[object] = [
            rank,
            window.window.depart.isoformat(),
            window.window.return_.isoformat(),
            float(total) if total is not None else None,
        ]
        for name in traveler_names:
            result = window.per_traveler.get(name)
            offer = result.best_offer if result else None
            if offer is None:
                row.append(None)
                row.append("—")
            else:
                row.append(float(offer.price))
                row.append(f"{offer.provider}/{offer.origin}")
        if total is not None and baseline is not None:
            row.append(float(baseline - total))
        else:
            row.append(None)
        row.append("yes" if window.complete else "no")
        row.append(", ".join(window.missing) if window.missing else "")
        ws.append(row)
    _autosize(ws)
    return ws


def _dashboard_tab(
    wb: Workbook, report: ComparisonReport, traveler_names: list[str]
) -> None:
    ws = wb.create_sheet("Dashboard")
    ranked = report.ranked()

    # Data block: one row per window with group total + per-traveler prices.
    ws.append(["Window", "Group total", *traveler_names])
    for window in ranked:
        total = window.group_total
        row: list[object] = [
            str(window.window),
            float(total) if total is not None else None,
        ]
        for name in traveler_names:
            result = window.per_traveler.get(name)
            offer = result.best_offer if result else None
            row.append(float(offer.price) if offer is not None else None)
        ws.append(row)

    n_rows = len(ranked)
    if n_rows == 0:
        return
    last_data_row = 1 + n_rows

    # Bar chart: group cost by window.
    bar = BarChart()
    bar.title = "Group cost by window"
    bar.type = "col"
    bar.y_axis.title = f"Group total ({report.currency})"
    bar.x_axis.title = "Window"
    data = Reference(ws, min_col=2, min_row=1, max_row=last_data_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 18
    ws.add_chart(bar, "H2")

    # Line chart: per-traveler cost trend across windows.
    line = LineChart()
    line.title = "Per-traveler cost trend"
    line.y_axis.title = f"Fare ({report.currency})"
    line.x_axis.title = "Window"
    tdata = Reference(
        ws,
        min_col=3,
        max_col=2 + len(traveler_names),
        min_row=1,
        max_row=last_data_row,
    )
    line.add_data(tdata, titles_from_data=True)
    line.set_categories(cats)
    line.height = 8
    line.width = 18
    ws.add_chart(line, "H20")


def export_report(
    report: ComparisonReport,
    traveler_names: list[str],
    out_path: str | Path,
) -> Path:
    """Write ``report`` to an ``.xlsx`` workbook and return its path."""
    wb = Workbook()
    _fares_tab(wb, report)
    _comparison_tab(wb, report, traveler_names)
    _dashboard_tab(wb, report, traveler_names)
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


__all__ = ["export_report"]
