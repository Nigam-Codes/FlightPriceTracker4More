"""Smoke test the .xlsx export (structure + charts present)."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from converge_flights.export import export_report
from converge_flights.models import (
    ComparisonReport,
    DateWindow,
    TravelerResult,
    WindowComparison,
)
from tests.conftest import make_offer

WINDOW = DateWindow(depart=date(2025, 9, 11), return_=date(2025, 9, 14))


def _report() -> ComparisonReport:
    comparison = WindowComparison(
        window=WINDOW,
        currency="USD",
        per_traveler={
            "Alex": TravelerResult(
                traveler="Alex",
                window=WINDOW,
                best_offer=make_offer(origin="JFK", price="280.00"),
                considered=3,
            ),
            "Sam": TravelerResult(
                traveler="Sam",
                window=WINDOW,
                best_offer=make_offer(origin="DTW", price="200.00"),
                considered=1,
            ),
        },
    )
    return ComparisonReport(currency="USD", windows=[comparison])


def test_export_creates_workbook(tmp_path: Path) -> None:
    out = tmp_path / "results.xlsx"
    written = export_report(_report(), ["Alex", "Sam"], out)
    assert written.exists()

    wb = load_workbook(written)
    assert set(wb.sheetnames) == {"Fares", "Comparison", "Dashboard"}

    dashboard = wb["Dashboard"]
    # Both charts should have been attached to the dashboard sheet.
    assert len(dashboard._charts) == 2  # noqa: SLF001 - test introspection

    comparison = wb["Comparison"]
    header = [c.value for c in next(comparison.iter_rows(max_row=1))]
    assert "Group total" in header
